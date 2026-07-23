"""Genera los archivos xlsx de los recibos llenando las plantillas del BGC-software.

Las coordenadas (celdas y filas de inicio) son las mismas que usaba el software
de escritorio original — vienen calcadas de `utils/recibo_generator_*.py` del
repo `BGC-software`. No se cambian aquí: la validación humana ya se hizo sobre
esos xlsx durante años.
"""

from __future__ import annotations

import io
from dataclasses import dataclass, field
from datetime import date
from importlib import resources
from typing import Any

from coop_core.utils.formato import format_full_name_for_excel, format_miles_colombian_int
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.worksheet.properties import PageSetupProperties

_MAX_FILAS = 6
_GASTO_POR_APORTE = 3000
_START_ROW = 9

_RECIBO_ID_CELL = "D4"
_FECHA_CELL = "I4"
_RECIBI_DE_CELL = "G6"


@dataclass(frozen=True)
class SocioBasico:
    nombres: str
    apellidos: str


@dataclass(frozen=True)
class LineaAporte:
    socio: SocioBasico
    monto: int
    saldo_anterior: int
    saldo_nuevo: int


@dataclass(frozen=True)
class LineaPago:
    socio: SocioBasico
    letra_id: int
    nro_cuota_inicio: int | str
    nro_cuota_fin: int | str
    total_cuotas_letra: int
    saldo_capital_antes: int
    abono_capital: int
    interes: int
    saldo_capital_despues: int
    mora: int = 0


@dataclass
class DatosRecibo:
    recibo_id: int
    fecha: date
    recibi_de: SocioBasico
    aportes: list[LineaAporte] = field(default_factory=list)
    pagos: list[LineaPago] = field(default_factory=list)
    num_aportes_cobrables: int | None = None
    monto_retiro: int | None = None
    socio_retiro: SocioBasico | None = None


@dataclass(frozen=True)
class CuotaLiquidacion:
    nro_cuota: int
    fecha_vencimiento: str
    valor_cuota: int
    interes_mes: int
    cuota_mensual: int
    saldo_capital: int
    fecha_pago: str = ""  # vacío en la liquidación original; con fecha en la actual


@dataclass
class DatosLiquidacion:
    letra_id: int
    capital: int
    interes: float
    n_cuotas: int
    fecha_inicio: date
    socios: list[SocioBasico]
    valor_cuota_base: int
    cuotas: list[CuotaLiquidacion]


# ── Utilidades comunes ────────────────────────────────────────────────────────


def _abrir_plantilla(subruta: str) -> Any:
    with resources.as_file(resources.files("coop_api.recibos.templates").joinpath(subruta)) as p:
        return load_workbook(p)


def _cabecera_estandar(ws: Any, recibo_id: int, fecha: date, recibi_de: SocioBasico) -> None:
    ws[_RECIBO_ID_CELL] = recibo_id
    ws[_FECHA_CELL] = fecha.strftime("%d/%m/%Y")
    ws[_RECIBI_DE_CELL] = f"{recibi_de.nombres} {recibi_de.apellidos}".upper()
    ws[_RECIBI_DE_CELL].alignment = Alignment(horizontal="center")


def _guardar(wb: Any) -> bytes:
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _forzar_horizontal(ws: Any) -> None:
    """Orientación horizontal + ajuste a una página de ancho, para que
    LibreOffice no corte el recibo al convertir a PDF."""
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)


def _cuota_display(pago: LineaPago) -> str:
    inicio, fin = pago.nro_cuota_inicio, pago.nro_cuota_fin
    if isinstance(inicio, str) and "ABONO" in inicio.upper():
        return "ABONO"
    if inicio == fin:
        return f"{inicio}/{pago.total_cuotas_letra}"
    return f"{inicio}-{fin}/{pago.total_cuotas_letra}"


# ── Generadores por tipo ──────────────────────────────────────────────────────


def generar_xlsx_aporte(datos: DatosRecibo) -> bytes:
    aportes = datos.aportes[:_MAX_FILAS]
    n = len(aportes)
    if n == 0:
        raise ValueError("Recibo de aportes sin aportes")

    wb = _abrir_plantilla(f"recibo_template_aporte/recibo_template_aporte{n}.xlsx")
    ws = wb.active
    _cabecera_estandar(ws, datos.recibo_id, datos.fecha, datos.recibi_de)

    total = 0
    for i, aporte in enumerate(aportes):
        row = _START_ROW + i
        nombre = format_full_name_for_excel(aporte.socio.nombres, aporte.socio.apellidos, max_length=24)
        ws[f"B{row}"] = nombre
        ws[f"F{row}"] = format_miles_colombian_int(aporte.saldo_anterior)
        ws[f"H{row}"] = format_miles_colombian_int(aporte.monto)
        ws[f"J{row}"] = format_miles_colombian_int(aporte.saldo_nuevo)
        total += aporte.monto

    row_total = _START_ROW + n
    row_admin = row_total + 2
    row_general = row_total + 3

    cobrables = datos.num_aportes_cobrables if datos.num_aportes_cobrables is not None else n
    gastos_admin = _GASTO_POR_APORTE * cobrables

    ws[f"H{row_total}"] = format_miles_colombian_int(total)
    ws[f"K{row_admin}"] = format_miles_colombian_int(gastos_admin)
    ws[f"K{row_general}"] = format_miles_colombian_int(total + gastos_admin)

    return _guardar(wb)


def generar_xlsx_pago(datos: DatosRecibo) -> bytes:
    pagos = datos.pagos[:_MAX_FILAS]
    n = len(pagos)
    if n == 0:
        raise ValueError("Recibo de pagos sin pagos")

    wb = _abrir_plantilla(f"recibo_template_pago/recibo_template_pago{n}.xlsx")
    ws = wb.active
    _cabecera_estandar(ws, datos.recibo_id, datos.fecha, datos.recibi_de)

    total_capital_interes = 0
    total_mora = 0
    for i, pago in enumerate(pagos):
        row = _START_ROW + i
        nombre = format_full_name_for_excel(pago.socio.nombres, pago.socio.apellidos, max_length=24)
        ws[f"B{row}"] = nombre
        ws[f"F{row}"] = pago.letra_id
        ws[f"G{row}"] = _cuota_display(pago)
        ws[f"H{row}"] = format_miles_colombian_int(pago.saldo_capital_antes)
        ws[f"I{row}"] = format_miles_colombian_int(pago.abono_capital)
        ws[f"J{row}"] = format_miles_colombian_int(pago.interes)
        ws[f"K{row}"] = format_miles_colombian_int(pago.saldo_capital_despues)
        total_capital_interes += pago.abono_capital + pago.interes
        total_mora += pago.mora

    row_total = _START_ROW + n
    row_admin = row_total + 2
    row_mora = row_admin + 1
    row_general = row_mora + 1

    ws[f"H{row_total}"] = format_miles_colombian_int(total_capital_interes)
    ws[f"K{row_admin}"] = format_miles_colombian_int(0)
    ws[f"K{row_mora}"] = format_miles_colombian_int(total_mora)
    ws[f"K{row_general}"] = format_miles_colombian_int(total_capital_interes + total_mora)

    return _guardar(wb)


def generar_xlsx_retiro(datos: DatosRecibo) -> bytes:
    if datos.socio_retiro is None or datos.monto_retiro is None:
        raise ValueError("Recibo de retiro necesita socio_retiro y monto_retiro")

    wb = _abrir_plantilla("recibo_template_retiro.xlsx")
    ws = wb.active
    socio = datos.socio_retiro

    ws["B6"] = datos.recibo_id
    ws["F6"] = format_miles_colombian_int(datos.monto_retiro)
    ws["F8"] = datos.fecha.strftime("%d/%m/%Y")
    ws["C14"] = f"DEVOLUCION PARCIAL DE APORTES DE {socio.nombres.upper()}"
    ws["C15"] = socio.apellidos.upper()
    ws["C18"] = f"{socio.nombres} {socio.apellidos}".upper()
    ws["C18"].alignment = Alignment(horizontal="center")

    # La plantilla de retiro viene en vertical y el PDF se corta al convertir.
    # Se fuerza horizontal + ajuste a una página de ancho, como los demás recibos.
    _forzar_horizontal(ws)

    return _guardar(wb)


def generar_xlsx_combinado(datos: DatosRecibo) -> bytes:
    aportes = datos.aportes[:_MAX_FILAS]
    pagos = datos.pagos[:_MAX_FILAS]
    x, y = len(aportes), len(pagos)
    if x == 0 or y == 0:
        raise ValueError("Recibo combinado necesita al menos un aporte y un pago")

    wb = _abrir_plantilla(f"recibo_template_combinado/recibo_template_combinado{x}_{y}.xlsx")
    ws = wb.active
    _cabecera_estandar(ws, datos.recibo_id, datos.fecha, datos.recibi_de)

    # Sección aportes
    total_aportes = 0
    for i, aporte in enumerate(aportes):
        row = _START_ROW + i
        nombre = format_full_name_for_excel(aporte.socio.nombres, aporte.socio.apellidos)
        ws[f"B{row}"] = nombre
        ws[f"F{row}"] = format_miles_colombian_int(aporte.saldo_anterior)
        ws[f"H{row}"] = format_miles_colombian_int(aporte.monto)
        ws[f"J{row}"] = format_miles_colombian_int(aporte.saldo_nuevo)
        total_aportes += aporte.monto

    row_total_aportes = _START_ROW + x
    ws[f"H{row_total_aportes}"] = format_miles_colombian_int(total_aportes)

    # Sección créditos (arranca 3 filas después del total de aportes)
    start_creditos = row_total_aportes + 3
    total_creditos = 0
    total_mora = 0
    for i, pago in enumerate(pagos):
        row = start_creditos + i
        nombre = format_full_name_for_excel(pago.socio.nombres, pago.socio.apellidos)
        ws[f"B{row}"] = nombre
        ws[f"F{row}"] = pago.letra_id
        ws[f"G{row}"] = _cuota_display(pago)
        ws[f"H{row}"] = format_miles_colombian_int(pago.saldo_capital_antes)
        ws[f"I{row}"] = format_miles_colombian_int(pago.abono_capital)
        ws[f"J{row}"] = format_miles_colombian_int(pago.interes)
        ws[f"K{row}"] = format_miles_colombian_int(pago.saldo_capital_despues)
        total_creditos += pago.abono_capital + pago.interes
        total_mora += pago.mora

    row_total_creditos = start_creditos + y
    row_admin = row_total_creditos + 2
    row_mora = row_admin + 1
    row_general = row_mora + 1

    cobrables = datos.num_aportes_cobrables if datos.num_aportes_cobrables is not None else x
    gastos_admin = _GASTO_POR_APORTE * cobrables

    ws[f"H{row_total_creditos}"] = format_miles_colombian_int(total_creditos)
    ws[f"K{row_admin}"] = format_miles_colombian_int(gastos_admin)
    ws[f"K{row_mora}"] = format_miles_colombian_int(total_mora)
    ws[f"K{row_general}"] = format_miles_colombian_int(
        total_aportes + total_creditos + gastos_admin + total_mora
    )

    return _guardar(wb)


# ── Liquidación de crédito ────────────────────────────────────────────────────

_LIQ_TABLA_START_ROW = 14
_LIQ_TABLA_HEADERS = [
    "Fecha",
    "Cuota",
    "Valor Cuota",
    "Intereses",
    "Total Mensual",
    "Saldo Capital",
    "Fecha Pago",
]
_LIQ_TABLA_COLUMNS = ["A", "B", "C", "D", "E", "F", "G"]
_LIQ_FIRMA_TESORERO = "Tesorero: ALVARO L. BURBANO GARCIA"


def generar_xlsx_liquidacion(datos: DatosLiquidacion) -> bytes:
    """Calca `utils/credit_liquidation_generator.py` del BGC-software.

    La tabla de amortización viene ya calculada (misma lógica de
    `coop_core.services.amortization`), aquí solo se vuelca a la plantilla.
    """
    from openpyxl.styles import Border, Font, Side

    borde = Border(
        left=Side(style="thin", color="FF000000"),
        right=Side(style="thin", color="FF000000"),
        top=Side(style="thin", color="FF000000"),
        bottom=Side(style="thin", color="FF000000"),
    )
    fuente_header = Font(bold=True)
    centro = Alignment(horizontal="center", vertical="center")

    wb = _abrir_plantilla("recibo_template_liquidacion.xlsx")
    ws = wb.active

    ws["B7"] = datos.letra_id
    ws["F7"] = format_miles_colombian_int(datos.capital)
    socios_nombres = [f"{s.nombres} {s.apellidos}".upper().strip() for s in datos.socios]
    ws["B9"] = ", Y/O ".join(socios_nombres)
    ws["A12"] = datos.fecha_inicio.strftime("%Y-%m-%d")
    ws["B12"] = datos.n_cuotas
    ws["C12"] = format_miles_colombian_int(datos.valor_cuota_base)
    ws["D12"] = f"{datos.interes * 100:.2f}%"
    ws["F12"] = format_miles_colombian_int(datos.capital)

    fila = _LIQ_TABLA_START_ROW
    for col_idx, header in enumerate(_LIQ_TABLA_HEADERS):
        celda = ws[f"{_LIQ_TABLA_COLUMNS[col_idx]}{fila}"]
        celda.value = header
        celda.font = fuente_header
        celda.alignment = centro
        celda.border = borde

    fila += 1
    for cuota in datos.cuotas:
        valores = [
            cuota.fecha_vencimiento,
            str(cuota.nro_cuota),
            format_miles_colombian_int(cuota.valor_cuota),
            format_miles_colombian_int(cuota.interes_mes),
            format_miles_colombian_int(cuota.cuota_mensual),
            format_miles_colombian_int(max(0, cuota.saldo_capital)),
            cuota.fecha_pago,
        ]
        for col_idx, valor in enumerate(valores):
            celda = ws[f"{_LIQ_TABLA_COLUMNS[col_idx]}{fila}"]
            celda.value = valor
            celda.alignment = centro
            celda.border = borde
        fila += 1

    for col in _LIQ_TABLA_COLUMNS:
        ws.column_dimensions[col].width = 15

    fila_firma = fila + 2
    ws.merge_cells(start_row=fila_firma, start_column=1, end_row=fila_firma, end_column=7)
    celda_firma = ws[f"A{fila_firma}"]
    celda_firma.value = _LIQ_FIRMA_TESORERO
    celda_firma.alignment = centro
    celda_firma.font = Font(bold=True, size=12)

    return _guardar(wb)


# ── Recibo de salario ─────────────────────────────────────────────────────────


@dataclass
class DatosSalario:
    recibo_id: int
    valor: int
    fecha: date
    mes: str  # en palabra, p.ej. "Junio"


def generar_xlsx_salario(datos: DatosSalario) -> bytes:
    """Llena la plantilla de salario. Solo cuatro celdas (las combinadas se
    llenan desde su celda superior-izquierda):
      B4 = número de recibo · G4 = valor · D6 = fecha · C12 = mes."""
    wb = _abrir_plantilla("recibo_template_salario.xlsx")
    ws = wb.active
    ws["B4"] = datos.recibo_id
    ws["G4"] = format_miles_colombian_int(datos.valor)
    ws["D6"] = datos.fecha.strftime("%d/%m/%Y")
    ws["C12"] = datos.mes
    return _guardar(wb)
