"""Tests del generador de xlsx: verifica que las celdas quedan bien llenas.

No depende de LibreOffice — abre el xlsx generado con openpyxl y valida.
"""

from __future__ import annotations

import io
from datetime import date

from coop_api.recibos.generador import (
    CuotaLiquidacion,
    DatosLiquidacion,
    DatosRecibo,
    DatosSalario,
    LineaAporte,
    LineaPago,
    SocioBasico,
    generar_xlsx_aporte,
    generar_xlsx_combinado,
    generar_xlsx_liquidacion,
    generar_xlsx_pago,
    generar_xlsx_retiro,
    generar_xlsx_salario,
)
from openpyxl import load_workbook


def _abrir(xlsx: bytes):
    wb = load_workbook(io.BytesIO(xlsx))
    return wb.active


def test_aporte_llena_cabecera_y_totales() -> None:
    datos = DatosRecibo(
        recibo_id=42,
        fecha=date(2026, 7, 22),
        recibi_de=SocioBasico(nombres="Jose David", apellidos="Ruano Burbano"),
        aportes=[
            LineaAporte(
                socio=SocioBasico(nombres="Jose David", apellidos="Ruano Burbano"),
                monto=200000,
                saldo_anterior=5380000,
                saldo_nuevo=5580000,
            ),
        ],
        num_aportes_cobrables=1,
    )
    ws = _abrir(generar_xlsx_aporte(datos))
    assert ws["D4"].value == 42
    assert ws["I4"].value == "22/07/2026"
    assert "JOSE DAVID RUANO BURBANO" in str(ws["G6"].value)
    # Fila del primer aporte
    assert "200.000" in str(ws["H9"].value)
    assert "5.580.000" in str(ws["J9"].value)
    # Total y admin
    assert "200.000" in str(ws["H10"].value)
    assert "3.000" in str(ws["K12"].value)  # gastos_admin = 3000 * 1 aporte cobrable


def test_retiro_llena_campos_de_devolucion() -> None:
    socio = SocioBasico(nombres="María Elena", apellidos="Gómez Ruiz")
    datos = DatosRecibo(
        recibo_id=7,
        fecha=date(2026, 7, 22),
        recibi_de=socio,
        socio_retiro=socio,
        monto_retiro=350000,
    )
    ws = _abrir(generar_xlsx_retiro(datos))
    assert ws["B6"].value == 7
    assert "350.000" in str(ws["F6"].value)
    assert ws["F8"].value == "22/07/2026"
    assert "MARÍA ELENA" in str(ws["C14"].value)
    assert str(ws["C15"].value) == "GÓMEZ RUIZ"
    assert str(ws["C18"].value) == "MARÍA ELENA GÓMEZ RUIZ"


def test_pago_llena_una_cuota() -> None:
    datos = DatosRecibo(
        recibo_id=100,
        fecha=date(2026, 7, 22),
        recibi_de=SocioBasico(nombres="Pedro", apellidos="Gómez"),
        pagos=[
            LineaPago(
                socio=SocioBasico(nombres="Pedro", apellidos="Gómez"),
                letra_id=451,
                nro_cuota_inicio=3,
                nro_cuota_fin=3,
                total_cuotas_letra=12,
                saldo_capital_antes=800000,
                abono_capital=70000,
                interes=8000,
                saldo_capital_despues=730000,
            ),
        ],
    )
    ws = _abrir(generar_xlsx_pago(datos))
    assert ws["D4"].value == 100
    assert ws["F9"].value == 451
    assert ws["G9"].value == "3/12"
    assert "800.000" in str(ws["H9"].value)
    assert "70.000" in str(ws["I9"].value)
    assert "8.000" in str(ws["J9"].value)


def test_combinado_usa_matriz_x_y() -> None:
    datos = DatosRecibo(
        recibo_id=200,
        fecha=date(2026, 7, 22),
        recibi_de=SocioBasico(nombres="Karoll", apellidos="Marcela"),
        aportes=[
            LineaAporte(
                socio=SocioBasico(nombres="Karoll", apellidos="Marcela"),
                monto=100000,
                saldo_anterior=500000,
                saldo_nuevo=600000,
            ),
        ],
        pagos=[
            LineaPago(
                socio=SocioBasico(nombres="Karoll", apellidos="Marcela"),
                letra_id=450,
                nro_cuota_inicio=2,
                nro_cuota_fin=3,
                total_cuotas_letra=10,
                saldo_capital_antes=900000,
                abono_capital=150000,
                interes=12000,
                saldo_capital_despues=750000,
            ),
        ],
        num_aportes_cobrables=1,
    )
    ws = _abrir(generar_xlsx_combinado(datos))
    assert ws["D4"].value == 200
    # Aporte en fila 9
    assert "100.000" in str(ws["H9"].value)
    # Pago 3 filas después del total de aportes (row_total=10, start_creditos=13)
    assert ws["F13"].value == 450
    assert ws["G13"].value == "2-3/10"


def test_liquidacion_llena_cabecera_y_tabla() -> None:
    datos = DatosLiquidacion(
        letra_id=77,
        capital=1200000,
        interes=0.01,
        n_cuotas=3,
        fecha_inicio=date(2026, 7, 22),
        socios=[SocioBasico(nombres="Pedro", apellidos="Gómez Ruiz")],
        valor_cuota_base=400000,
        cuotas=[
            CuotaLiquidacion(
                nro_cuota=1,
                fecha_vencimiento="2026-08-22",
                valor_cuota=400000,
                interes_mes=12000,
                cuota_mensual=412000,
                saldo_capital=800000,
            ),
            CuotaLiquidacion(
                nro_cuota=2,
                fecha_vencimiento="2026-09-22",
                valor_cuota=400000,
                interes_mes=8000,
                cuota_mensual=408000,
                saldo_capital=400000,
            ),
            CuotaLiquidacion(
                nro_cuota=3,
                fecha_vencimiento="2026-10-22",
                valor_cuota=400000,
                interes_mes=4000,
                cuota_mensual=404000,
                saldo_capital=0,
            ),
        ],
    )
    ws = _abrir(generar_xlsx_liquidacion(datos))
    assert ws["B7"].value == 77
    assert "1.200.000" in str(ws["F7"].value)
    assert "PEDRO GÓMEZ RUIZ" in str(ws["B9"].value)
    assert ws["B12"].value == 3
    assert ws["D12"].value == "1.00%"
    # Encabezados de la tabla en fila 14
    assert ws["A14"].value == "Fecha"
    assert ws["G14"].value == "Fecha Pago"
    # Primera cuota en fila 15
    assert ws["A15"].value == "2026-08-22"
    assert ws["B15"].value == "1"
    assert "12.000" in str(ws["D15"].value)


def test_salario_llena_las_cuatro_celdas() -> None:
    datos = DatosSalario(recibo_id=77, valor=1500000, fecha=date(2026, 6, 15), mes="Junio")
    ws = _abrir(generar_xlsx_salario(datos))
    assert ws["B4"].value == 77
    assert "1.500.000" in str(ws["G4"].value)
    assert ws["D6"].value == "15/06/2026"
    assert ws["C12"].value == "Junio"
